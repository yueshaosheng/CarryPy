#!/usr/bin/env python3
# [目标机执行] 每包功能冒烟测试 (不只 import, 真正跑一遍核心功能):
#   用法: python3 selfcheck_pkgs.py                       # 测所有已安装的已知包
#         python3 selfcheck_pkgs.py 包1 包2              # 只测指定包 (pypi 名, 可带版本约束)
#         python3 selfcheck_pkgs.py --output-dir DIR      # 测试产物保存到 DIR
#         python3 selfcheck_pkgs.py --output-dir DIR 包1  # 组合使用
#   - 未安装的包标记 SKIP (无参模式) / FAIL (点名模式, 点名了就必须在)
#   - 没有内置测试用例的包回退为 import 验证
#   - 有 FAIL 时退出码非 0
import importlib
import importlib.util
import os
import re
import sys
import tempfile

# pypi 包名 -> import 模块名 (与 install_addon.sh 的映射保持一致)
ALIAS = {
    "scikit-learn": "sklearn",
    "pillow": "PIL",
    "opencv-python": "cv2",
    "pyyaml": "yaml",
    "pyside2": "PySide2",
    "pyside6": "PySide6",
    "pyqt5": "PyQt5",
    "pyqt6": "PyQt6",
}


def _file_ok(path):
    assert os.path.getsize(path) > 0, f"输出文件为空: {path}"


def test_numpy(tmp):
    import numpy as np
    a = np.random.rand(20, 20)
    inv = np.linalg.inv(a)
    assert np.allclose(a @ inv, np.eye(20)), "矩阵求逆回乘校验失败"


def test_matplotlib(tmp):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots()
    ax.add_patch(plt.Circle((0.5, 0.5), 0.3, fill=False))
    ax.set_aspect("equal")
    out = os.path.join(tmp, "circle.png")
    fig.savefig(out)
    plt.close(fig)
    _file_ok(out)


def test_pandas(tmp):
    import pandas as pd
    df = pd.DataFrame({"k": list("aabb"), "v": [1, 2, 3, 4]})
    g = df.groupby("k")["v"].sum()
    assert g["a"] == 3 and g["b"] == 7, "groupby 结果不对"
    out = os.path.join(tmp, "df.csv")
    df.to_csv(out, index=False)
    back = pd.read_csv(out)
    assert back.equals(df), "csv 写读回环不一致"


def test_scipy(tmp):
    import numpy as np
    from scipy import stats, integrate
    p = stats.ttest_1samp(np.random.randn(50), 0).pvalue
    assert 0 <= p <= 1
    area, _ = integrate.quad(lambda x: x * x, 0, 3)
    assert abs(area - 9) < 1e-6, f"积分结果 {area} != 9"


def test_seaborn(tmp):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np
    import pandas as pd
    import seaborn as sns
    df = pd.DataFrame({"x": np.random.randn(50), "y": np.random.randn(50)})
    sns.scatterplot(data=df, x="x", y="y")
    out = os.path.join(tmp, "scatter.png")
    plt.savefig(out)
    plt.close("all")
    _file_ok(out)


def test_plotly(tmp):
    import plotly.graph_objects as go
    import plotly.io as pio
    fig = go.Figure(go.Scatter(x=[1, 2, 3], y=[4, 5, 6]))
    back = pio.from_json(fig.to_json())
    assert list(back.data[0].y) == [4, 5, 6], "to_json 回环数据不一致"


def test_pillow(tmp):
    from PIL import Image, ImageDraw
    img = Image.new("RGB", (64, 64), "white")
    ImageDraw.Draw(img).ellipse((8, 8, 56, 56), outline="red", width=3)
    out = os.path.join(tmp, "circle.png")
    img.save(out)
    back = Image.open(out)
    assert back.size == (64, 64), "图像读回尺寸不对"


def test_openpyxl(tmp):
    from openpyxl import Workbook, load_workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["name", "value"])
    ws.append(["alpha", 1])
    ws.append(["beta", 2])
    out = os.path.join(tmp, "test.xlsx")
    wb.save(out)
    wb2 = load_workbook(out, read_only=True)
    ws2 = wb2.active
    rows = list(ws2.iter_rows(values_only=True))
    wb2.close()
    assert rows[0] == ("name", "value"), f"表头不对: {rows[0]}"
    assert rows[1] == ("alpha", 1), f"数据不对: {rows[1]}"


def test_sklearn(tmp):
    import numpy as np
    from sklearn.linear_model import LogisticRegression
    rng = np.random.RandomState(0)
    X = rng.randn(100, 4)
    y = (X[:, 0] + X[:, 1] > 0).astype(int)
    score = LogisticRegression().fit(X, y).score(X, y)
    assert score > 0.7, f"LogisticRegression score 过低: {score}"


def _test_qt(binding, tmp):
    # 无显示器环境也能跑: 强制 offscreen 平台插件
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    mod = importlib.import_module(binding + ".QtGui")
    app = mod.QGuiApplication.instance() or mod.QGuiApplication([])
    img = mod.QImage(64, 64, mod.QImage.Format.Format_RGB32) \
        if hasattr(mod.QImage, "Format") and hasattr(mod.QImage.Format, "Format_RGB32") \
        else mod.QImage(64, 64, mod.QImage.Format_RGB32)
    p = mod.QPainter(img)
    p.fillRect(8, 8, 48, 48, mod.QColor("steelblue"))
    p.end()
    out = os.path.join(tmp, "qt.png")
    assert img.save(out), "QImage.save 失败"
    _file_ok(out)


TESTS = {
    "numpy": ("numpy", test_numpy),
    "matplotlib": ("matplotlib", test_matplotlib),
    "pandas": ("pandas", test_pandas),
    "scipy": ("scipy", test_scipy),
    "seaborn": ("seaborn", test_seaborn),
    "plotly": ("plotly", test_plotly),
    "pillow": ("PIL", test_pillow),
    "openpyxl": ("openpyxl", test_openpyxl),
    "scikit-learn": ("sklearn", test_sklearn),
    "pyside2": ("PySide2", lambda tmp: _test_qt("PySide2", tmp)),
    "pyside6": ("PySide6", lambda tmp: _test_qt("PySide6", tmp)),
    "pyqt5": ("PyQt5", lambda tmp: _test_qt("PyQt5", tmp)),
    "pyqt6": ("PyQt6", lambda tmp: _test_qt("PyQt6", tmp)),
}


def norm(pkg):
    """去掉版本约束, 统一小写: 'Scikit-Learn==1.3' -> 'scikit-learn'"""
    return re.split(r"[=<>!~\[]", pkg, 1)[0].strip().lower()


def main(argv):
    # 解析 --output-dir 参数
    output_dir = None
    raw_args = []
    i = 0
    while i < len(argv):
        if argv[i] == "--output-dir" and i + 1 < len(argv):
            output_dir = argv[i + 1]
            i += 2
        else:
            raw_args.append(argv[i])
            i += 1

    explicit = [norm(a) for a in raw_args if norm(a)]
    names = explicit or list(TESTS)

    # 有 output_dir 时测试产物保存到那里, 否则用临时目录 (测完自动清理)
    use_tmp = output_dir is None
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    n_pass = n_fail = 0
    for name in names:
        modname = TESTS[name][0] if name in TESTS else ALIAS.get(name, name.replace("-", "_"))
        if importlib.util.find_spec(modname) is None:
            if explicit:                       # 点名模式: 点名了就必须已安装
                print(f"  [FAIL] {name}: 未安装 (找不到模块 {modname})")
                n_fail += 1
            # 默认模式: 跳过未安装的包, 不显示
            continue
        try:
            mod = importlib.import_module(modname)
            ver = getattr(mod, "__version__", "?")
            if name in TESTS:
                if use_tmp:
                    with tempfile.TemporaryDirectory() as tmp:
                        TESTS[name][1](tmp)
                else:
                    TESTS[name][1](output_dir)
                print(f"  [PASS] {name} {ver}: 功能测试通过")
            else:                              # 无内置用例: 回退 import 验证
                print(f"  [PASS] {name} {ver}: import 验证通过 (无内置功能测试)")
            n_pass += 1
        except Exception as e:
            print(f"  [FAIL] {name}: {type(e).__name__}: {e}")
            n_fail += 1

    print(f"  ---- 冒烟测试汇总: {n_pass} PASS, {n_fail} FAIL ----")
    if output_dir:
        print(f"  测试产物已保存到: {output_dir}")
    return 1 if n_fail else 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
